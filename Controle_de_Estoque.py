import win32com.client as win32
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File
import pandas as pd
from io import BytesIO
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, PatternFill, Side
from copy import copy

# Função para autenticar no SharePoint
def autenticar_sharepoint(sharepoint_url, username, password):
    auth_context = AuthenticationContext(sharepoint_url)
    
    if auth_context.acquire_token_for_user(username, password):
        return ClientContext(sharepoint_url, auth_context)

    return None  # Retorna None se a autenticação falhar



def enviar_email_outlook(ctx, file_relative_url, nome, pn, quantidade_atual, estoque_minimo, estoque_maximo, tipo_inventario, codigo_manual=None, obs=None):
    try:
        # Extrair dados da aba Emails
        response = File.open_binary(ctx, file_relative_url)
        bytes_file = BytesIO()
        bytes_file.write(response.content)
        bytes_file.seek(0)
        excel_data = pd.read_excel(bytes_file, sheet_name=None, engine="openpyxl")

        if "Emails" in excel_data:
            emails_df = excel_data["Emails"]
            destinatarios = emails_df["Destinatários"].dropna().tolist()
            copia = emails_df["Cópia"].dropna().tolist()
        else:
            messagebox.showerror("Erro", "A aba 'Emails' não foi encontrada na planilha.")
            return

        # Formatar o corpo do e-mail para Ferramentas
        if tipo_inventario == "Ferramentas":
            corpo_email = f"""
            <p>Prezado(a),</p>
            <p>O sistema de controle de estoque chegou no ponto de reposição para a seguinte ferramenta:</p>
            <ul>
                <li><b>Nome:</b> {nome}</li>
                <li><b>PN:</b> {pn}</li>
                <li><b>Quantidade Atual:</b> {quantidade_atual}</li>
                <li><b>Estoque Mínimo:</b> {estoque_minimo}</li>
                <li><b>Estoque Máximo:</b> {estoque_maximo}</li>
                <li><b>Quantidade Sugerida para Compra:</b> {estoque_maximo - quantidade_atual}</li>
            </ul>
            <p>*Quantidade sugerida para compra é a quantidade necessária para atingir o estoque máximo.</p>
            <p>**Essa solicitação foi gerada automaticamente pelo sistema.</p>
            <p>Att,<br>Sistema de Controle de Estoque</p>
            """
        else:  # Caso seja um Consumível
            corpo_email = f"""
            <p>Prezado(a),</p>
            <p>O sistema de controle de estoque chegou no ponto de reposição para o seguinte consumível:</p>
            <ul>
                <li><b>Nome:</b> {nome}</li>
                <li><b>Código do Manual:</b> {codigo_manual if codigo_manual else 'N/A'}</li>
                <li><b>Quantidade Atual:</b> {quantidade_atual}</li>
                <li><b>Estoque Mínimo:</b> {estoque_minimo}</li>
                <li><b>Estoque Máximo:</b> {estoque_maximo}</li>
                <li><b>Quantidade Sugerida para Compra:</b> {estoque_maximo - quantidade_atual}</li>
                <li><b>Observação:</b> {obs if obs else 'Nenhuma'}</li>
            </ul>
            <p>*Quantidade sugerida para compra é a quantidade necessária para atingir o estoque máximo.</p>
            <p>**Essa solicitação foi gerada automaticamente pelo sistema.</p>
            <p>Att,<br>Sistema de Controle de Estoque</p>
            """

        # Envio do e-mail
        outlook = win32.Dispatch("outlook.application")
        mail = outlook.CreateItem(0)
        mail.To = "; ".join(destinatarios)
        mail.CC = "; ".join(copia)
        mail.Subject = f"Solicitação de Compra"
        mail.HTMLBody = corpo_email
        mail.Send()

        print(f"E-mail enviado com sucesso para {tipo_inventario}!")
    except Exception as e:
        print(f"Erro ao enviar e-mail: {e}")

# Função para extrair dados da planilha
def extrair_dados(ctx, file_relative_url):
    try:
        response = File.open_binary(ctx, file_relative_url)
        bytes_file = BytesIO()
        bytes_file.write(response.content)
        bytes_file.seek(0)
        df = pd.read_excel(bytes_file, sheet_name=None, engine="openpyxl")
        return df
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao extrair dados: Email ou Senha incorretos.")
        return None

# Função para salvar a planilha no SharePoint
def salvar_planilha(ctx, excel_data, file_relative_url):
    fonte_poppins = Font(name="Poppins", size=10)
    alinhamento_central = Alignment(horizontal="center", vertical="center")
    try:
        # Baixa o arquivo original do SharePoint
        response = File.open_binary(ctx, file_relative_url)
        bytes_file = BytesIO()
        bytes_file.write(response.content)
        bytes_file.seek(0)

        # Carrega o workbook preservando formatação
        wb = load_workbook(bytes_file)

        for sheet_name, data in excel_data.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Mantém largura das colunas
                col_widths = {col: ws.column_dimensions[col].width for col in ws.column_dimensions}

                # Mantém altura das linhas
                row_heights = {row: ws.row_dimensions[row].height for row in ws.row_dimensions}

                # Mantém estilos de célula (bordas, fontes, alinhamento, fundo)
                styles = {}
                for row in ws.iter_rows():
                    for cell in row:
                        styles[cell.coordinate] = {
                            "font": copy(cell.font),
                            "border": copy(cell.border),
                            "alignment": copy(cell.alignment),
                            "fill": copy(cell.fill)
                        }

                # Apaga os dados antigos sem remover o cabeçalho
                ws.delete_rows(2, ws.max_row)

                # Escreve os novos dados mantendo estilos
                for row_index, row in enumerate(data.itertuples(index=False), start=2):
                    for col_index, value in enumerate(row, start=1):
                        cell = ws.cell(row=row_index, column=col_index, value=value)

                        # Aplica formatação original se existir
                        coord = cell.coordinate
                        if coord in styles:
                            cell.font = styles[coord]["font"]
                            cell.border = styles[coord]["border"]
                            cell.alignment = styles[coord]["alignment"]
                            cell.fill = styles[coord]["fill"]
                        else:
                            # Se não houver formatação original, usa Poppins 10
                            cell.font = fonte_poppins
                            cell.alignment = alinhamento_central

                # Restaura largura das colunas e altura das linhas
                for col, width in col_widths.items():
                    ws.column_dimensions[col].width = width
                for row, height in row_heights.items():
                    ws.row_dimensions[row].height = height

            else:
                # Cria nova aba caso não exista
                ws = wb.create_sheet(sheet_name)
                for col_index, col_name in enumerate(data.columns, start=1):
                    ws.cell(row=1, column=col_index, value=col_name)
                for row_index, row in enumerate(data.itertuples(index=False), start=2):
                    for col_index, value in enumerate(row, start=1):
                        ws.cell(row=row_index, column=col_index, value=value)

        # Salva de volta para o SharePoint
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        folder_url = file_relative_url.rsplit("/", 1)[0]
        file_name = file_relative_url.rsplit("/", 1)[1]
        ctx.web.get_folder_by_server_relative_url(folder_url).upload_file(file_name, output.getvalue()).execute_query()

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao salvar a planilha: {e}")

# Função para registrar operações na aba "Registros"
def registrar_operacao(nome_mecanico, item, quantidade, operacao, ctx, file_relative_url, tipo_inventario, prateleira=None):
    try:
        response = File.open_binary(ctx, file_relative_url)
        bytes_file = BytesIO()
        bytes_file.write(response.content)
        bytes_file.seek(0)
        excel_data = pd.read_excel(bytes_file, sheet_name=None, engine="openpyxl")

        # Definir nome da aba de registros e colunas corretas
        aba_registros = "Registros"
        col_item = "PN" if tipo_inventario == "Ferramentas" else "Consumível"

        # Se a aba de registros não existir, criar
        if aba_registros in excel_data:
            registros_df = excel_data[aba_registros].dropna(how='all')
        else:
            registros_df = pd.DataFrame(columns=["Nome", col_item, "Quantidade", "Operação", "Data"])

        # Adicionar coluna de Prateleira se for ferramenta
        if tipo_inventario == "Ferramentas":
            if "Prateleira" not in registros_df.columns:
                registros_df["Prateleira"] = None  # Adiciona coluna caso esteja faltando

        # Garantir que as colunas existam e estejam no formato correto
        for col in ["Nome", col_item, "Quantidade", "Operação", "Data"]:
            if col not in registros_df.columns:
                registros_df[col] = None

        # Converter quantidades para int para evitar erros
        registros_df["Quantidade"] = pd.to_numeric(registros_df["Quantidade"], errors='coerce').fillna(0).astype(int)

        # Criar nova entrada corretamente
        nova_linha = {
            "Nome": nome_mecanico,  # Nome do mecânico
            col_item: item,          # Nome do item (PN ou Consumível)
            "Quantidade": quantidade,
            "Operação": operacao,
            "Data": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        nova_linha_df = pd.DataFrame([nova_linha])

        # Adicionar ao dataframe de registros e resetar índice
        registros_df = pd.concat([registros_df, nova_linha_df], ignore_index=True).reset_index(drop=True)

        # Salvar a planilha com os novos registros
        excel_data[aba_registros] = registros_df
        salvar_planilha(ctx, excel_data, file_relative_url)

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao registrar operação: {e}")


# Função para atualizar a quantidade de itens
def atualizar_quantidade(pn, quantidade, nome, operacao, ctx, file_relative_urls, tipo_inventario, prateleira=None):
    try:
        file_relative_url = file_relative_urls[tipo_inventario]
        response = File.open_binary(ctx, file_relative_url)
        bytes_file = BytesIO()
        bytes_file.write(response.content)
        bytes_file.seek(0)
        excel_data = pd.read_excel(bytes_file, sheet_name=None, engine="openpyxl")

        # Definir a aba correta e os nomes das colunas
        if tipo_inventario == "Ferramentas":
            aba = "Prateleira Principal" if prateleira == "Prateleira Principal" else "Prateleira Secundária"
            col_pn = "PN" if prateleira == "Prateleira Principal" else "Ferramenta"  # Na Secundária, é "Ferramenta"
            col_nome = "Ferramenta"
            col_quantidade = "Quantidade"
            col_estoque_min = "Estoque Mínimo"
            col_estoque_max = "Estoque Máximo"
            col_situacao = "Situação"
        else:
            aba = "Estoque"
            col_pn = "Nome"  # Para Consumíveis, a referência é o nome do item
            col_nome = "Nome"
            col_quantidade = "Quantidade"
            col_estoque_min = "Estoque Mínimo"
            col_estoque_max = "Estoque Máximo"
            col_situacao = "Situação"

        # Verifica se a aba existe
        if aba not in excel_data:
            messagebox.showerror("Erro", f"A aba '{aba}' não foi encontrada na planilha de {tipo_inventario}.")
            return

        estoque_df = excel_data[aba]

        # Garante que as colunas existem e converte para int antes de usar
        for col in [col_quantidade, col_estoque_min, col_estoque_max]:
            if col in estoque_df.columns:
                estoque_df[col] = pd.to_numeric(estoque_df[col], errors="coerce").fillna(0).astype(int)
            else:
                messagebox.showerror("Erro", f"A coluna '{col}' não foi encontrada na planilha de {tipo_inventario}.")
                return

        # Verifica se o item existe antes de prosseguir
        if pn not in estoque_df[col_pn].values:
            print(f"ERRO: Item '{pn}' não encontrado na aba '{aba}'.")  # Depuração
            messagebox.showerror("Erro", f"O item '{pn}' não foi encontrado no estoque da {aba}.")
            return  # Impede que o código continue e cause erros

        # Obtém o índice correto do item na planilha
        index = estoque_df[estoque_df[col_pn] == pn].index[0]
        current_quantity = estoque_df.at[index, col_quantidade]
        estoque_minimo = estoque_df.at[index, col_estoque_min]
        estoque_maximo = estoque_df.at[index, col_estoque_max]
        ferramenta = estoque_df.at[index, col_nome]
        situacao_atual = estoque_df.at[index, col_situacao]

        # Lógica de entrada e saída de estoque
        if operacao == "Saída":
            if current_quantity < quantidade:
                messagebox.showerror("Erro", "Quantidade insuficiente em estoque.")
                return
            estoque_df.at[index, col_quantidade] -= quantidade

            if estoque_df.at[index, col_quantidade] <= estoque_minimo and situacao_atual == "Em estoque":
                enviar_email_outlook(
                    ctx,
                    file_relative_url,
                    nome=ferramenta,
                    pn=pn,
                    quantidade_atual=estoque_df.at[index, col_quantidade],
                    estoque_minimo=estoque_minimo,
                    estoque_maximo=estoque_maximo,
                    tipo_inventario=tipo_inventario,
                    codigo_manual=estoque_df.at[index, "Código do Manual"] if "Código do Manual" in estoque_df.columns else None,
                    obs=estoque_df.at[index, "Obs"] if "Obs" in estoque_df.columns else None
                )
                estoque_df.at[index, col_situacao] = "Enviado para compras"
        elif operacao == "Entrada":
            estoque_df.at[index, col_quantidade] += quantidade
            estoque_df.at[index, col_situacao] = "Em estoque"

        # Salvar as alterações na planilha
        excel_data[aba] = estoque_df
        salvar_planilha(ctx, excel_data, file_relative_url)

        # Registrar operação com a prateleira escolhida
        registrar_operacao(nome, pn, quantidade, operacao, ctx, file_relative_url, tipo_inventario, prateleira)

        messagebox.showinfo("Sucesso", "Quantidade atualizada com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao atualizar a quantidade: {e}")


def configurar_tabela(frame_table, df):
    # Configurando estilo para o Treeview
    style = ttk.Style()

    # Definição das cores conforme a formatação fornecida
    bh_gray = "#D9D9D9"  # Cinza BH para headers destacados
    bh_light_gray = "#F2F2F2"  # Cinza claro para colunas destacadas
    border_gray = "#B3B3B3"  # Cinza para as bordas das linhas

    # Criar um novo layout específico para os headers
    style.layout("Treeview.Heading", [
        ("Treeheading.cell", {"sticky": "nswe"}),
        ("Treeheading.border", {"sticky": "nswe", "children": [
            ("Treeheading.padding", {"sticky": "nswe", "children": [
                ("Treeheading.image", {"side": "right", "sticky": ""}),
                ("Treeheading.text", {"sticky": "we"})
            ]})
        ]})
    ])

    # Configuração do cabeçalho (headers)
    style.configure("Treeview.Heading",
                    font=("Poppins SemiBold", 12),
                    background=bh_gray,  # Cinza BH para destaque
                    foreground="black",  # Texto PRETO para melhor visibilidade
                    relief="flat",
                    padding=10)

    # Configuração do corpo da tabela
    style.configure("Treeview",
                    font=("Poppins", 12),  # Texto da tabela entre 10-14pt
                    rowheight=30,
                    background="white",
                    foreground="black",  # Garante que todo o texto esteja em preto
                    fieldbackground="white",
                    borderwidth=1,  # Bordas de 1pt
                    relief="flat")

    style.map("Treeview",
              background=[("selected", "#d0d0d0")],  # Azul claro quando selecionado
              foreground=[("selected", "black")])  # Texto preto mesmo quando selecionado

    # Criando o Treeview
    tree = ttk.Treeview(frame_table, columns=list(df.columns), show="headings")
    tree.pack(side="left", expand=True, fill="both")

    # Configurando colunas e cabeçalhos
    for col in df.columns:
        tree.heading(col, text=col, anchor="center")
        tree.column(col, width=150, anchor="center")

    # Adicionando linhas alternadas com destaque
    for index, row in df.iterrows():
        tags = "oddrow" if index % 2 == 0 else "evenrow"
        tree.insert("", "end", values=list(row), tags=(tags,))

    tree.tag_configure("oddrow", background="white", foreground="black")  # Fundo branco, texto preto
    tree.tag_configure("evenrow", background=bh_light_gray, foreground="black")  # Cinza claro para destaque, texto preto

    # Barra de rolagem
    scrollbar = ttk.Scrollbar(frame_table, orient="vertical", command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    scrollbar.pack(side="right", fill="y")

    return tree


# Interface gráfica
def iniciar_interface():

    def login():
        username = username_entry.get().strip()
        password = password_entry.get().strip()

        if not username or not password:
            messagebox.showerror("Erro", "Os campos de Username e Password não podem estar vazios.")
            return

        global ctx
        ctx = autenticar_sharepoint(sharepoint_url, username, password)

        # Verifica se a autenticação falhou
        if ctx is None:
            messagebox.showerror("Erro", "Falha na autenticação. Verifique suas credenciais e tente novamente.")
            return

        # Testa se a conexão é válida acessando um recurso do SharePoint
        try:
            ctx.web.get().execute_query()
        except Exception as e:
            messagebox.showerror("Erro", "Falha na autenticação. Credenciais inválidas.")
            return

        # Se a autenticação for bem-sucedida, prossegue para a próxima tela
        exibir_tela_escolha()


    def exibir_tela_escolha():
        # Limpa a janela principal
        for widget in root.winfo_children():
            widget.destroy()

        # Cabeçalho
        tk.Label(root, text="Escolha o Inventário", font=("Helvetica", 20, "bold"), pady=20).pack()

        # Criar frame para os botões
        frame = tk.Frame(root)
        frame.pack(pady=50)

        # Botão para acessar o inventário de ferramentas
        ferramentas_button = tk.Button(
            frame, text="Inventário de Ferramentas", font=("Helvetica", 14), width=25,
            command=lambda: mostrar_dados("Ferramentas")
        )
        ferramentas_button.grid(row=0, column=0, padx=20, pady=10)

        # Botão para acessar o inventário de consumíveis
        consumiveis_button = tk.Button(
            frame, text="Inventário de Consumíveis", font=("Helvetica", 14), width=25,
            command=lambda: mostrar_dados("Consumíveis")
        )
        consumiveis_button.grid(row=1, column=0, padx=20, pady=10)

        # Botão de logout
        logout_button = tk.Button(
            frame, text="Sair", font=("Helvetica", 14), width=25, command=root.destroy
        )
        logout_button.grid(row=2, column=0, padx=20, pady=10)



    def mostrar_dados(tipo_inventario):
        global df, search_entry, tree
        planilha = extrair_dados(ctx, file_relative_urls[tipo_inventario])

        if planilha:
            if tipo_inventario == "Ferramentas":
                abas_ferramentas = ["Prateleira Principal", "Prateleira Secundária"]
                df_list = []

                for aba in abas_ferramentas:
                    if aba in planilha:
                        df_temp = planilha[aba]
                        df_temp["Origem"] = aba  # Adiciona uma coluna para identificar a origem
                        df_list.append(df_temp)

                if df_list:
                    df = pd.concat(df_list, ignore_index=True)
                else:
                    messagebox.showerror("Erro", "Nenhuma aba de ferramentas encontrada na planilha.")
                    return

                colunas_desejadas = ["Ferramenta", "PN", "Quantidade", "Prateleira", "Situação", "Origem"]
                df = df[colunas_desejadas]

            else:
                aba_principal = "Estoque"
                if aba_principal in planilha:
                    df = planilha[aba_principal]
                else:
                    messagebox.showerror("Erro", f"A aba '{aba_principal}' não foi encontrada na planilha de {tipo_inventario}.")
                    return

                colunas_desejadas = ["Nome", "Código do Manual", "Quantidade", "Situação", "Obs", "Validade"]
                df = df[colunas_desejadas]

            # Limpa a tela antes de carregar os dados
            for widget in root.winfo_children():
                widget.destroy()

            root.geometry("1100x600")

            # Cabeçalho
            tk.Label(root, text=f"Inventário de {tipo_inventario}", font=("Helvetica", 20, "bold"), pady=20).pack()

            # Barra de pesquisa
            search_frame = tk.Frame(root)
            search_frame.pack(pady=10)
            tk.Label(search_frame, text="Pesquisar:", font=("Helvetica", 14)).grid(row=0, column=0, padx=10)
            search_entry = tk.Entry(search_frame, width=30, font=("Helvetica", 14))
            search_entry.grid(row=0, column=1, padx=10)
            search_button = tk.Button(search_frame, text="Pesquisar", command=pesquisar, font=("Helvetica", 12))
            search_button.grid(row=0, column=2, padx=10)

            # Configurando a tabela
            frame_table = tk.Frame(root)
            frame_table.pack(expand=True, fill="both")
            tree = configurar_tabela(frame_table, df)

            # Botões
            frame_buttons = tk.Frame(root)
            frame_buttons.pack(pady=20)

            adicionar_button = tk.Button(
                frame_buttons,
                text="Adicionar Item",
                command=lambda: adicionar_item(tipo_inventario, ctx, file_relative_urls),
                width=15,
                font=("Helvetica", 14)
            )
            adicionar_button.grid(row=0, column=0, padx=10)

            remover_button = tk.Button(
                frame_buttons,
                text="Remover Item",
                command=lambda: remover_item(tipo_inventario, ctx, file_relative_urls),
                width=15,
                font=("Helvetica", 14)
            )
            remover_button.grid(row=0, column=1, padx=10)

            adicionar_novo_button = tk.Button(
                frame_buttons,
                text="Adicionar Novo Item",
                command=lambda: adicionar_item_novo(tipo_inventario, ctx, file_relative_urls),
                width=20,
                font=("Helvetica", 14)
            )
            adicionar_novo_button.grid(row=0, column=2, padx=10)

            voltar_button = tk.Button(
                frame_buttons,
                text="Voltar",
                command=exibir_tela_escolha,
                width=15,
                font=("Helvetica", 14)
            )
            voltar_button.grid(row=0, column=3, padx=10)

            sair_button = tk.Button(
                frame_buttons,
                text="Sair",
                command=root.destroy,
                width=15,
                font=("Helvetica", 14)
            )
            sair_button.grid(row=0, column=4, padx=10)


    def pesquisar():
        termo = search_entry.get().strip()

        if not termo:  # Se o campo de pesquisa estiver vazio, mostrar todos os dados
            atualizar_tabela(df)
            return

        # Verifica o tipo de inventário e ajusta a pesquisa na coluna correta
        if "PN" in df.columns or "Ferramenta" in df.columns:  # Se for Ferramentas
            df_filtrado = df[
                df["PN"].astype(str).str.contains(termo, na=False, case=False) |
                df["Ferramenta"].astype(str).str.contains(termo, na=False, case=False)
            ]
        elif "Nome" in df.columns:  # Se for Consumíveis
            df_filtrado = df[df["Nome"].astype(str).str.contains(termo, na=False, case=False)]
        else:
            messagebox.showerror("Erro", "Não foi possível realizar a pesquisa.")
            return

        atualizar_tabela(df_filtrado)  # Atualiza a exibição da tabela com os resultados da pesquisa


    def atualizar_tabela(dataframe):
        for row in tree.get_children():
            tree.delete(row)

        for index, row in dataframe.iterrows():
            tags = "oddrow" if index % 2 == 0 else "evenrow"
            tree.insert("", "end", values=list(row), tags=(tags,))

        tree.tag_configure("oddrow", background="white")
        tree.tag_configure("evenrow", background="lightblue")

    def adicionar_item(tipo_inventario, ctx, file_relative_urls):
        def salvar():
            nome = nome_entry.get()
            quantidade = quantidade_entry.get()

            # Se for ferramentas, definimos a prateleira
            if tipo_inventario == "Ferramentas":
                prateleira = prateleira_var.get()
                if prateleira == "Prateleira Principal":
                    pn = pn_entry.get().strip().upper()  # Usuário digita o PN
                else:
                    pn = nome_combobox.get().strip()  # Usuário escolhe pelo Nome da Ferramenta
            else:
                prateleira = None  # Consumíveis não têm prateleira
                pn = nome_combobox.get().strip()

            if not nome or not pn or not quantidade:
                messagebox.showerror("Erro", "Todos os campos devem ser preenchidos.")
                return

            try:
                quantidade = int(quantidade)
                atualizar_quantidade(pn, quantidade, nome, "Entrada", ctx, file_relative_urls, tipo_inventario, prateleira)
                adicionar_window.destroy()
                mostrar_dados(tipo_inventario)  # Atualiza a tela
            except ValueError:
                messagebox.showerror("Erro", "A quantidade deve ser um número inteiro.")

        # Criando a janela para adicionar item
        adicionar_window = tk.Toplevel()
        adicionar_window.title("Adicionar Item")
        adicionar_window.geometry("500x300")

        tk.Label(adicionar_window, text="Nome do Mecânico:", font=("Helvetica", 12)).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        nome_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
        nome_entry.grid(row=0, column=1, padx=10, pady=5)

        if tipo_inventario == "Ferramentas":
            # Escolha da Prateleira
            tk.Label(adicionar_window, text="Prateleira:", font=("Helvetica", 12)).grid(row=1, column=0, padx=10, pady=5, sticky="w")
            prateleira_var = tk.StringVar(value="Prateleira Principal")
            prateleira_combobox = ttk.Combobox(adicionar_window, textvariable=prateleira_var, values=["Prateleira Principal", "Prateleira Secundária"], font=("Helvetica", 12), width=28)
            prateleira_combobox.grid(row=1, column=1, padx=10, pady=5)

            def atualizar_interface_prateleira(event):
                if prateleira_var.get() == "Prateleira Principal":
                    pn_entry.grid(row=2, column=1, padx=10, pady=5)
                    nome_combobox.grid_forget()
                else:
                    pn_entry.grid_forget()
                    nome_combobox.grid(row=2, column=1, padx=10, pady=5)

            prateleira_combobox.bind("<<ComboboxSelected>>", atualizar_interface_prateleira)

            # Entrada de PN manual para Prateleira Principal
            tk.Label(adicionar_window, text="PN:", font=("Helvetica", 12)).grid(row=2, column=0, padx=10, pady=5, sticky="w")
            pn_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
            pn_entry.grid(row=2, column=1, padx=10, pady=5)

            # Lista dinâmica para Prateleira Secundária (baseada na coluna "Ferramenta")
            file_relative_url = file_relative_urls[tipo_inventario]
            response = File.open_binary(ctx, file_relative_url)
            bytes_file = BytesIO()
            bytes_file.write(response.content)
            bytes_file.seek(0)
            excel_data = pd.read_excel(bytes_file, sheet_name=None, engine="openpyxl")

            nomes_disponiveis = []
            if "Prateleira Secundária" in excel_data:
                estoque_df = excel_data["Prateleira Secundária"]
                nomes_disponiveis = estoque_df["Ferramenta"].dropna().tolist()  # Pegamos os nomes das Ferramentas

            nome_combobox = ttk.Combobox(adicionar_window, values=nomes_disponiveis, width=28, font=("Helvetica", 12))
            nome_combobox.grid(row=2, column=1, padx=10, pady=5)
            nome_combobox.grid_forget()  # Esconde inicialmente

        else:  # Para Consumíveis
            # Seleção dinâmica dos consumíveis
            file_relative_url = file_relative_urls[tipo_inventario]
            response = File.open_binary(ctx, file_relative_url)
            bytes_file = BytesIO()
            bytes_file.write(response.content)
            bytes_file.seek(0)
            excel_data = pd.read_excel(bytes_file, sheet_name=None, engine="openpyxl")

            nomes_disponiveis = []
            if "Estoque" in excel_data:
                estoque_df = excel_data["Estoque"]
                nomes_disponiveis = estoque_df["Nome"].dropna().tolist()

            tk.Label(adicionar_window, text="Nome do Item:", font=("Helvetica", 12)).grid(row=1, column=0, padx=10, pady=5, sticky="w")
            nome_combobox = ttk.Combobox(adicionar_window, values=nomes_disponiveis, width=28, font=("Helvetica", 12))
            nome_combobox.grid(row=1, column=1, padx=10, pady=5)

        # Entrada da Quantidade
        tk.Label(adicionar_window, text="Quantidade:", font=("Helvetica", 12)).grid(row=3, column=0, padx=10, pady=5, sticky="w")
        quantidade_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
        quantidade_entry.grid(row=3, column=1, padx=10, pady=5)

        salvar_button = tk.Button(adicionar_window, text="Salvar", command=salvar, font=("Helvetica", 12))
        salvar_button.grid(row=4, column=0, columnspan=2, pady=20)


    def remover_item(tipo_inventario, ctx, file_relative_urls):
        def salvar():
            nome = nome_entry.get()
            quantidade = quantidade_entry.get()

            # Se for Ferramentas, definir a prateleira
            if tipo_inventario == "Ferramentas":
                prateleira = prateleira_var.get()
                if prateleira == "Prateleira Principal":
                    pn = pn_entry.get().strip().upper()  # Usuário digita o PN
                else:
                    pn = nome_combobox.get().strip()  # Usuário escolhe pelo Nome da Ferramenta
            else:
                prateleira = None  # Consumíveis não têm prateleira
                pn = nome_combobox.get().strip()

            if not nome or not pn or not quantidade:
                messagebox.showerror("Erro", "Todos os campos devem ser preenchidos.")
                return

            try:
                quantidade = int(quantidade)
                atualizar_quantidade(pn, quantidade, nome, "Saída", ctx, file_relative_urls, tipo_inventario, prateleira)
                remover_window.destroy()
                mostrar_dados(tipo_inventario)  # Atualiza a tela após a remoção
            except ValueError:
                messagebox.showerror("Erro", "A quantidade deve ser um número inteiro.")

        # Criando janela para remover item
        remover_window = tk.Toplevel()
        remover_window.title("Remover Item")
        remover_window.geometry("500x300")

        tk.Label(remover_window, text="Nome do Mecânico:", font=("Helvetica", 12)).grid(row=0, column=0, padx=10, pady=5)
        nome_entry = tk.Entry(remover_window, width=30, font=("Helvetica", 12))
        nome_entry.grid(row=0, column=1, padx=10, pady=5)

        if tipo_inventario == "Ferramentas":
            # Escolha da Prateleira
            tk.Label(remover_window, text="Prateleira:", font=("Helvetica", 12)).grid(row=1, column=0, padx=10, pady=5)
            prateleira_var = tk.StringVar(value="Prateleira Principal")
            prateleira_combobox = ttk.Combobox(remover_window, textvariable=prateleira_var, values=["Prateleira Principal", "Prateleira Secundária"], font=("Helvetica", 12), width=28)
            prateleira_combobox.grid(row=1, column=1, padx=10, pady=5)

            def atualizar_interface_prateleira(event):
                if prateleira_var.get() == "Prateleira Principal":
                    pn_entry.grid(row=2, column=1, padx=10, pady=5)
                    nome_combobox.grid_forget()
                else:
                    pn_entry.grid_forget()
                    nome_combobox.grid(row=2, column=1, padx=10, pady=5)

            prateleira_combobox.bind("<<ComboboxSelected>>", atualizar_interface_prateleira)

            # Entrada de PN manual para Prateleira Principal
            tk.Label(remover_window, text="PN:", font=("Helvetica", 12)).grid(row=2, column=0, padx=10, pady=5)
            pn_entry = tk.Entry(remover_window, width=30, font=("Helvetica", 12))
            pn_entry.grid(row=2, column=1, padx=10, pady=5)

            # Lista dinâmica para Prateleira Secundária (baseada na coluna "Ferramenta")
            file_relative_url = file_relative_urls[tipo_inventario]
            response = File.open_binary(ctx, file_relative_url)
            bytes_file = BytesIO()
            bytes_file.write(response.content)
            bytes_file.seek(0)
            excel_data = pd.read_excel(bytes_file, sheet_name=None, engine="openpyxl")

            nomes_disponiveis = []
            if "Prateleira Secundária" in excel_data:
                estoque_df = excel_data["Prateleira Secundária"]
                nomes_disponiveis = estoque_df["Ferramenta"].dropna().tolist()  # Pegamos os nomes das Ferramentas

            nome_combobox = ttk.Combobox(remover_window, values=nomes_disponiveis, width=28, font=("Helvetica", 12))
            nome_combobox.grid(row=2, column=1, padx=10, pady=5)
            nome_combobox.grid_forget()  # Esconde inicialmente

        else:  # Para Consumíveis
            # Seleção dinâmica dos consumíveis
            file_relative_url = file_relative_urls[tipo_inventario]
            response = File.open_binary(ctx, file_relative_url)
            bytes_file = BytesIO()
            bytes_file.write(response.content)
            bytes_file.seek(0)
            excel_data = pd.read_excel(bytes_file, sheet_name=None, engine="openpyxl")

            nomes_disponiveis = []
            if "Estoque" in excel_data:
                estoque_df = excel_data["Estoque"]
                nomes_disponiveis = estoque_df["Nome"].dropna().tolist()

            tk.Label(remover_window, text="Nome do Item:", font=("Helvetica", 12)).grid(row=1, column=0, padx=10, pady=5)
            nome_combobox = ttk.Combobox(remover_window, values=nomes_disponiveis, width=28, font=("Helvetica", 12))
            nome_combobox.grid(row=1, column=1, padx=10, pady=5)

        # Entrada da Quantidade
        tk.Label(remover_window, text="Quantidade:", font=("Helvetica", 12)).grid(row=3, column=0, padx=10, pady=5)
        quantidade_entry = tk.Entry(remover_window, width=30, font=("Helvetica", 12))
        quantidade_entry.grid(row=3, column=1, padx=10, pady=5)

        salvar_button = tk.Button(remover_window, text="Salvar", command=salvar, font=("Helvetica", 12))
        salvar_button.grid(row=4, column=0, columnspan=2, pady=20)


    
    def adicionar_item_novo(tipo_inventario, ctx, file_relative_urls):
        def salvar():
            nome_mecanico = nome_mecanico_entry.get().strip()  # Nome do mecânico para registro
            nome = nome_entry.get().strip()
            quantidade = quantidade_entry.get().strip()
            estoque_minimo = estoque_minimo_entry.get().strip()
            estoque_maximo = estoque_maximo_entry.get().strip()
            situacao = "Em estoque"  # Sempre começa como "Em estoque"

            if tipo_inventario == "Ferramentas":
                pn = pn_entry.get().strip().upper()  # Sempre em maiúsculas
                local = local_var.get()  # Obtém se é Prateleira Principal ou Secundária
                prateleira = prateleira_entry.get().strip()  # Número da prateleira
                ponto_reposicao = ponto_reposicao_entry.get().strip()

                if not nome or not pn or not quantidade or not estoque_minimo or not estoque_maximo or not prateleira or not ponto_reposicao or not nome_mecanico:
                    messagebox.showerror("Erro", "Todos os campos devem ser preenchidos.")
                    return
            else:  # Para Consumíveis
                pn = None  # Consumíveis não têm PN
                codigo_manual = codigo_manual_entry.get().strip() or "N/A"
                obs = obs_entry.get().strip() or "Nenhuma"
                validade = validade_entry.get().strip() or "Não há"

                if not nome or not quantidade or not estoque_minimo or not estoque_maximo or not nome_mecanico:
                    messagebox.showerror("Erro", "Todos os campos devem ser preenchidos.")
                    return

            file_relative_url = file_relative_urls[tipo_inventario]
            response = File.open_binary(ctx, file_relative_url)
            bytes_file = BytesIO()
            bytes_file.write(response.content)
            bytes_file.seek(0)
            excel_data = pd.read_excel(bytes_file, sheet_name=None, engine="openpyxl")

            # Definir a aba correta e os nomes das colunas
            if tipo_inventario == "Ferramentas":
                aba = "Prateleira Principal" if local == "Prateleira Principal" else "Prateleira Secundária"
                col_pn = "PN" if local == "Prateleira Principal" else "Ferramenta"
                col_nome = "Ferramenta"
            else:
                aba = "Estoque"
                col_pn = "Nome"
                col_nome = "Nome"

            if aba not in excel_data:
                messagebox.showerror("Erro", f"A aba '{aba}' não foi encontrada na planilha.")
                return

            estoque_df = excel_data[aba]

            # Verificação de duplicação apenas para Ferramentas
            if tipo_inventario == "Ferramentas" and col_pn in estoque_df.columns and pn in estoque_df[col_pn].values:
                messagebox.showerror("Erro", f"O item {pn} já existe no estoque. Utilize a opção de adicionar itens.")
                return

            # Criando a nova linha do item
            if tipo_inventario == "Ferramentas":
                nova_linha = pd.DataFrame([{
                    "Ferramenta": nome,
                    "PN": pn,
                    "Quantidade": int(quantidade),
                    "Estoque Mínimo": int(estoque_minimo),
                    "Estoque Máximo": int(estoque_maximo),
                    "Ponto de reposição": int(ponto_reposicao),
                    "Prateleira": prateleira,
                    "Situação": situacao
                }])
            else:
                nova_linha = pd.DataFrame([{
                    "Nome": nome,
                    "Código do Manual": codigo_manual,
                    "Quantidade": int(quantidade),
                    "Situação": situacao,
                    "Estoque Mínimo": int(estoque_minimo),
                    "Estoque Máximo": int(estoque_maximo),
                    "Obs": obs,
                    "Validade": validade
                }])

            estoque_df = pd.concat([estoque_df, nova_linha], ignore_index=True)
            excel_data[aba] = estoque_df

            salvar_planilha(ctx, excel_data, file_relative_url)

            # Registrar operação corretamente, incluindo o nome do mecânico
            registrar_operacao(nome_mecanico, pn if tipo_inventario == "Ferramentas" else nome, quantidade, "Entrada", ctx, file_relative_url, tipo_inventario)

            messagebox.showinfo("Sucesso", "Novo item adicionado com sucesso!")
            adicionar_window.destroy()
            mostrar_dados(tipo_inventario)

        # Criando a janela do formulário
        adicionar_window = tk.Toplevel()
        adicionar_window.title("Adicionar Novo Item")
        adicionar_window.geometry("500x500")

        tk.Label(adicionar_window, text="Nome do Mecânico:", font=("Helvetica", 12)).grid(row=0, column=0, padx=10, pady=5)
        nome_mecanico_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
        nome_mecanico_entry.grid(row=0, column=1, padx=10, pady=5)

        tk.Label(adicionar_window, text="Nome do Item:", font=("Helvetica", 12)).grid(row=1, column=0, padx=10, pady=5)
        nome_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
        nome_entry.grid(row=1, column=1, padx=10, pady=5)

        if tipo_inventario == "Ferramentas":
            tk.Label(adicionar_window, text="Local:", font=("Helvetica", 12)).grid(row=2, column=0, padx=10, pady=5)
            local_var = tk.StringVar(value="Prateleira Principal")
            local_combobox = ttk.Combobox(adicionar_window, textvariable=local_var, values=["Prateleira Principal", "Prateleira Secundária"], font=("Helvetica", 12), width=28)
            local_combobox.grid(row=2, column=1, padx=10, pady=5)

            tk.Label(adicionar_window, text="PN:", font=("Helvetica", 12)).grid(row=3, column=0, padx=10, pady=5)
            pn_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
            pn_entry.grid(row=3, column=1, padx=10, pady=5)

            tk.Label(adicionar_window, text="Prateleira:", font=("Helvetica", 12)).grid(row=4, column=0, padx=10, pady=5)
            prateleira_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
            prateleira_entry.grid(row=4, column=1, padx=10, pady=5)

            tk.Label(adicionar_window, text="Ponto de Reposição:", font=("Helvetica", 12)).grid(row=5, column=0, padx=10, pady=5)
            ponto_reposicao_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
            ponto_reposicao_entry.grid(row=5, column=1, padx=10, pady=5)

        else:
            tk.Label(adicionar_window, text="Nome do Consumível:", font=("Helvetica", 12)).grid(row=1, column=0, padx=10, pady=5)
            nome_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
            nome_entry.grid(row=1, column=1, padx=10, pady=5)

            tk.Label(adicionar_window, text="Código do Manual (opcional):", font=("Helvetica", 12)).grid(row=2, column=0, padx=10, pady=5)
            codigo_manual_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
            codigo_manual_entry.grid(row=2, column=1, padx=10, pady=5)

            tk.Label(adicionar_window, text="Observação (opcional):", font=("Helvetica", 12)).grid(row=3, column=0, padx=10, pady=5)
            obs_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
            obs_entry.grid(row=3, column=1, padx=10, pady=5)

            tk.Label(adicionar_window, text="Validade (opcional):", font=("Helvetica", 12)).grid(row=4, column=0, padx=10, pady=5)
            validade_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
            validade_entry.grid(row=4, column=1, padx=10, pady=5)

        # Campos comuns para ambos os inventários
        tk.Label(adicionar_window, text="Quantidade:", font=("Helvetica", 12)).grid(row=6, column=0, padx=10, pady=5)
        quantidade_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
        quantidade_entry.grid(row=6, column=1, padx=10, pady=5)

        tk.Label(adicionar_window, text="Estoque Mínimo:", font=("Helvetica", 12)).grid(row=7, column=0, padx=10, pady=5)
        estoque_minimo_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
        estoque_minimo_entry.grid(row=7, column=1, padx=10, pady=5)

        tk.Label(adicionar_window, text="Estoque Máximo:", font=("Helvetica", 12)).grid(row=8, column=0, padx=10, pady=5)
        estoque_maximo_entry = tk.Entry(adicionar_window, width=30, font=("Helvetica", 12))
        estoque_maximo_entry.grid(row=8, column=1, padx=10, pady=5)

        salvar_button = tk.Button(adicionar_window, text="Salvar", command=salvar, font=("Helvetica", 12))
        salvar_button.grid(row=9, column=0, columnspan=2, pady=20)


    sharepoint_url = "url da empresa"
    # Dicionário contendo os caminhos das planilhas no SharePoint
    # Caso ocorra mudança de pastas, mudar apenas o file_relative_urls.
    # NÃO MUDAR OS NOMES "Ferramentas" E "Consumíveis" QUE ESTÃO NAS ASPAS ABAIXO
    file_relative_urls = {
        "Ferramentas": "local da planilha de inventário das ferramentas",
        "Consumíveis": "local da planilha de inventário dos consumíveis"
    }

    global ctx
    ctx = None

    root = tk.Tk()
    root.title("Sistema de Estoque")
    root.geometry("1000x600")

    header = tk.Label(root, text="Sistema de Estoque", font=("Helvetica", 20, "bold"), pady=20)
    header.pack()

    frame_login = tk.Frame(root)
    frame_login.pack(pady=30)
    tk.Label(frame_login, text="Username:", font=("Helvetica", 14)).grid(row=0, column=0, padx=10, pady=10)
    username_entry = tk.Entry(frame_login, width=30, font=("Helvetica", 14))
    username_entry.grid(row=0, column=1, padx=10, pady=10)
    tk.Label(frame_login, text="Password:", font=("Helvetica", 14)).grid(row=1, column=0, padx=10, pady=10)
    password_entry = tk.Entry(frame_login, width=30, show="*", font=("Helvetica", 14))
    password_entry.grid(row=1, column=1, padx=10, pady=10)

    frame_buttons = tk.Frame(root)
    frame_buttons.pack(pady=20)
    login_button = tk.Button(frame_buttons, text="Login", command=login, width=15, font=("Helvetica", 14))
    login_button.grid(row=0, column=0, padx=20, pady=10)
    sair_button = tk.Button(frame_buttons, text="Sair", command=root.destroy, width=15, font=("Helvetica", 14))
    sair_button.grid(row=0, column=1, padx=20, pady=10)
    footer = tk.Label(root, text="Baker Hughes", font=("Helvetica", 20, "bold"), fg="Dark Green")
    footer.pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    iniciar_interface()
